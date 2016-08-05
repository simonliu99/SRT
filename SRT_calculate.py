#Simon Liu
#07/20/16
#Quarknet 2016
#For use with JHU SRT and other similar .rad file outputs

import os
import openpyxl
import math
curDir = os.getcwd()

class gX(object):
    Name = ''
    Spec = []
    Bins = []
    Obs = 0
    Vlsr = 0
    Angle = 0
    Source = ''
    MinFreq = 0.0
    R = 0.0
    V = 0.0

def makegX(name):
    hD1,hD2,s = openFile(''.join((name,'.rad')))
    g = gX()
    g.Name = name
    g.Spec = s.split()
    g.Vlsr = hD1.split()[15]
    g.Obs = int(hD1.split()[3])
    g.Source = hD1.split()[-1]
    histDet = hD2.split()
    bins = makeBins(histDet, g.Spec)
    g.Bins = bins
    print hD1.split()[21]
    g.Angle = int(hD1.split()[21][1:])
    return g

def addgX(n,o,v,so,s,b):
    g = gX()
    g.Name = n
    g.Spec = s.split()
    g.Vlsr = v
    g.Obs = o
    g.Source = so
    g.Bins = b.split()
    g.Angle = int(so[1:])
    return g

def makeBins(b,s):
    minBin = float(b[1])
    maxBin = float(b[3])
    binWidth = float(b[5])
    curBin = minBin
    bins = []
    while bins.__len__() != s.__len__():
        bins.append(curBin)
        curBin += binWidth
    return bins

def openFile(fN):
    with open(fN) as f:
        fL = [line.strip('\n') for line in f]
    fileLen = file_len(fN)
    return fL[fileLen-4],fL[fileLen-3],fL[fileLen-1]

def file_len(fname):
    with open(fname) as f:
        for i, l in enumerate(f):
            pass
    return i + 1

def output(gL, dN):
    book = openpyxl.Workbook()
    for i in gL:
        sh = book.create_sheet(i.Name)
        for j, item in enumerate(i.Bins):
            sh.cell(row=j+1,column=1).value = float(item)
        for k, item in enumerate(i.Spec):
            sh.cell(row=k+1,column=2).value = float(item)
    sh = book.create_sheet('Overall')
    sh.cell(row=1,column=1).value = 'Bins'
    for h, item in enumerate(gL):
        sh.cell(row=1,column=h+2).value = item.Name
    for i, item in enumerate(gL[0].Bins):
            sh.cell(row=i+2,column=1).value = float(item)
    for j, item in enumerate(gL):
        for k, item2 in enumerate(item.Spec):
            sh.cell(row=k+2,column=j+2).value = float(item2)
    book.remove_sheet(book.get_sheet_by_name('Sheet'))
    book.save(''.join((dN,'.xlsx')))

def outOb(gL, dN):
    f = open(''.join((dN,'.txt')),'w')
    for i, item in enumerate(gL):
        f.write(''.join((gL[i].Name,'\n')))
        f.write(''.join((str(gL[i].Obs),'\n')))
        f.write(''.join((gL[i].Vlsr,'\n')))
        f.write(''.join((gL[i].Source,'\n')))
        f.write(''.join((' '.join(gL[i].Spec),'\n')))
        f.write(''.join((' '.join(map(str, gL[i].Bins)),'\n')))


def FNF():
    fR = os.listdir(os.getcwd())
    fF = []
    for f in fR:
        if f[f.__len__()-4:f.__len__()] == '.rad':
            fF.append(f[0:f.__len__()-4])
    return fF

def main():
    dN = raw_input('Directory: ')
    dJ = os.path.join(curDir, dN)
    os.chdir(dJ)
    fN = FNF()
    gL = []
    for i in fN:
        gXX = makegX(i)
        gL.append(gXX)
    print '1. .xlsx'
    print '2. .txt'
    print '3. Both .xlsx and .txt'
    op = input('Enter choice: ')
    if op == 1:
        output(gL,dN)
    elif op == 2:
        outOb(gL,dN)
    else:
        output(gL,dN)
        outOb(gL,dN)

def calc():
    vo = 220
    Ro = 245000000000000000
    wo = vo/Ro
    fo = 1420.405751
    c = 299792.458
    f = float(raw_input('Lowest Frequency: '))
    vlsr = float(raw_input('VLSR: '))
    v = -vlsr-(((f-fo)*c)/fo)
    print 'Maximum velocity = ', v
    gamma = input('Gamma Angle: ')
    sinGamma = math.sin(math.radians(gamma))
    vGalRot = v + 220*sinGamma
    print 'G',gamma,': ',vGalRot

def rotX():
    dN = raw_input('Directory: ')
    dJ = os.path.join(curDir, dN)
    os.chdir(dJ)
    with open(''.join((dN,'.txt'))) as f:
        fL = [line.strip('\n') for line in f]
    gL = []
    for i in range(0,fL.__len__()/6):
        j = 6 * i
        gXX = addgX(fL[j],fL[j+1],fL[j+2],fL[j+3],fL[j+4],fL[j+5])
        gL.append(gXX)
    rotCalc(gL,dN)

def minFreqTxt():
    dN = raw_input('Directory: ')
    dJ = os.path.join(curDir, dN)
    os.chdir(dJ)
    fN = FNF()
    mX = []
    for i in fN:
        mF = raw_input(''.join(('Enter minimum frequency for ',i,': ')))
        mX.append(mF)
    f = open(''.join((dN,'_minFreq.txt')),'w')
    for j, item in enumerate(mX):
        f.write(' '.join((fN[j],item,'\n')))

def minFreq(g,dN):
    fN = ''.join((dN,'_minFreq.txt'))
    with open(fN) as f:
        fL = [line.strip('\n') for line in f]
    freq = []
    for i, item in enumerate(fL):
        print item
        freq.append(item.split())
    for j in g:
        for k in freq:
            if k[0] == j.Name:
                j.MinFreq = float(k[1])

def rotCalc(g,dN):
    vo = 220
    Ro = 245000000000000000
    wo = vo/Ro
    fo = 1420.405751
    c = 299792.458
    minFreq(g,dN)
    for i in g:
        v = -(float(i.Vlsr))-(((float(i.MinFreq)-fo)*c)/fo)
        sinGamma = math.sin(math.radians(i.Angle))
        i.V = v + 220*sinGamma
        i.R = 7.94*sinGamma
        print i.Source, i.R, i.V
    book = openpyxl.Workbook()
    sh = book.create_sheet('Galaxy Rotation Curve')
    sh.cell(row=1,column=1).value = 'Radius (kpc)'
    sh.cell(row=1,column=2).value = 'Velocity (km/s)'
    for j, item in enumerate(g):
        sh.cell(row=j+2,column=1).value = item.R
        sh.cell(row=j+2,column=2).value = item.V
    book.remove_sheet(book.get_sheet_by_name('Sheet'))
    book.save(''.join((dN,'_GalRot.xlsx')))

def menu():
    print ''
    print 'Menu:'
    print '1. Process raw data files into .xlsx and .txt'
    print '2. Calculate galactic rotation velocity'
    print '3. Create minimum frequency text file'
    print '4. Create galactic rotation curve'
    print '5. Quit'
    choice = input('Enter choice: ')
    if choice == 1:
        main()
    elif choice == 2:
        calc()
    elif choice == 3:
        minFreqTxt()
    elif choice == 4:
        rotX()
    elif choice == 5:
        return True

q = False
while q != True:
    q = menu()
